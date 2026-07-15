import openpyxl
import re
from pathlib import Path
import json
import os
from datetime import datetime
import sys
from tools import *

# определяем основную директории проекта
BASE_DIR = Path(__file__).resolve().parent

# добавляем импорт из соседней папки
PATH_ROOT_DIR = BASE_DIR.parent 

sys.path.append(str(PATH_ROOT_DIR))

from splitter.main import find_file_by_regex


# определяем основные директории
SOURCE_DIR = BASE_DIR / 'source'
READY_DIR = BASE_DIR / 'ready'
CACHE_DIR = BASE_DIR / "cache"

# получаем файлы для дальнейшей обработки
path_admission = Path(find_file_by_regex(SOURCE_DIR, r"^Списки_поступающих_бакалавриат.*\.xlsx$"))
path_statement = Path(find_file_by_regex(SOURCE_DIR, r"^Ведомость.*\.xlsx$"))




def collect_data_statement(log_mode: bool) -> dict:
    """Функция собирает данные о баллах студентов с ведомостей.

    Args:
        log_mode (bool): Нужно ли выводить информацию о работе программы в консоль.

    Returns:
        dict: Словарик с собранными данными.
    """
    print("\nСобираю данные с ведомостей\n")

    wb = openpyxl.load_workbook(path_statement)

    # необходимые паттерны для поиска данных в шапке ведомости
    pattern_major = "Направление подготовки (специальность)"
    pattern_proile = "Профиль"
    pattern_test_name = "Предмет"
    patterns = [pattern_major, pattern_proile, pattern_test_name]

    expression =  "Спортивная подготовка по виду спорта. Тренерско-преподавательская деятельность в образовании"

    targets = ["B5", "B6", "B7"]

    # карта соотвествий
    value_mapping = {
        "Избранный вид спорта(Профессиональное испытание)": "избранный вид спорта",
        "Общая физическая подготовленность(Профессиональное испытание)": "офп"
    }

    # загружаем данные студентов
    data = smart_file_open()
    
    for ws in wb.worksheets:

        # делаем проверку, что нужные нам клеточки существуют
        for target, pattern in zip(targets, patterns):
            if ws[target].value != pattern:
                raise Exception("Нарушена структура в заголовке.")
        
        # получаем данные из шапки
        student_major = ws["D5"].value
        sport = ws["D6"].value # в ведомости указанг как профиль
        test_name = ws["D7"].value # название экзамена (предмет)

        # проверка на направление для которого должен быть указан sport
        if (sport and student_major != expression) or (student_major == expression and not sport):
            input(f"На листе {ws.title}, проблемы с видом спорта или направлением.")

        # обработка перменных
        student_major = render_student_major(sport, student_major) # обрабатываем направление подготовки
        test_name = value_mapping.get(test_name.strip()) # переводим название экзамена в нужный формат

        # если mode False ищем когда начнуться фамилии, если True уже нашли
        mode = False
        student_fio = str()

        # проходим по всем строчкам в файле
        current_row = 12
        while current_row <= ws.max_row:
            row = ws[current_row]

            # находим строку где начинаются данные
            if str(row[0].value) == "1" and str(row[1].value) != "2":
                mode = True

            # если начались нужные данные, записываем их
            if mode == True:
                student_fio = row[1].value
                test_points = row[7].value

            
            # если закончились фамилии, этот лист больше не нужно просматривать
            if student_fio == None:
                break
            
            # если мы нашли данные о студентах
            if mode == True:
                # делаем fio с маленькой буквы для удобства а баллы типом int
                student_fio = student_fio.lower()
                test_points = int(test_points)
                
                # вставляем данные в словарик
                insert_data_to_dict(
                    data=data,
                    student_fio=student_fio,
                    student_major=student_major,
                    test_name=test_name,
                    test_points=test_points,
                                    )
                
            if log_mode:
                print(f"student_fio -> {student_fio}, test_points -> {test_points}, test_name -> {test_name}")
                    
            current_row += 1

    with open(CACHE_DIR / "json.json", mode="w", encoding="utf-8") as f:
        json.dump(data, f,indent=4, ensure_ascii=False)
    
    print("\nСобрал все необходимые данные.")
    print("--------------------------------")

    return data


def admission_lists_add_points(data: dict, log_mode: bool) -> None:
    """Функция вставляет собранные значения из ведомостей в файл вступительные списки

    Args:
        data (dict): Данные которые нужно вставить.
        log_mode (bool): Нужно ли выводит информацию о работе программы в консоль.
    """

    print("\nВставляю данные во вступительные списки.\n")

    wd = openpyxl.load_workbook(path_admission)
    ws = wd.active

    inserted_data = list()

    pattern_fio: str = "фио"
    pattern_profile: str = "профиль"
    pattern_test_name: list = [
        "избранный вид спорта",
        "офп",
    ]

    dynamics_headers = [None, None]

    # проходим по всем строчкам в файле
    current_row = 1
    while current_row <= ws.max_row:
        row = ws[current_row]
        
        if row[0].value == "№":

            # сюда будем записывать изменяемые номера столбцов
            uuid_list = list()

            # сначала получаем не изменяемые названия столбцов
            uuid_dict = read_headers(
                row,
                pattern_fio, 
                pattern_profile,
                pattern_test_name[0],
                pattern_test_name[1]
                )

            # получаем индексы нужных нам колонок
            fio_id = uuid_dict[pattern_fio]
            profile_id = uuid_dict[pattern_profile]

            # списком с id столбцов, первый id ИВС второй ОФП
            dynamics_headers = [
                uuid_dict.get(pattern_test_name[0]),
                uuid_dict.get(pattern_test_name[1])
                                ]
        
        # важная проверка, чтобы не выполнять код, если таблица закончилась
        if row[0].value and not str(row[0].value).isdigit():
            pass

        # если нашли хотя бы один нужный заголовок
        elif dynamics_headers[0] or dynamics_headers[1]:

            
            
            # получаем необходимы ключи из таблицы
            cell_fio = row[fio_id].value
            cell_fio = cell_fio.lower() if isinstance(cell_fio, str) else None
            cell_profile = row[profile_id].value

            # получаем данные для заполнения из словарика
            for test_name, test_id in zip(pattern_test_name, dynamics_headers):

                # если в test_id есть значение, заполняем столбец под данным id
                if test_id:

                    # если такие данные есть в словаре, получаем их
                    input_data = get_data_from_dict(
                                    data=data,
                                    student_fio=cell_fio,
                                    student_major=cell_profile,
                                    test_name=test_name,
                                )
                    
                    row[test_id].value = input_data

                    if input_data:
                        inserted_data.append(cell_fio)
                    if input_data and log_mode:
                        print(f"ROW -> {current_row}, Фамилия -> {cell_fio}, баллы -> {input_data}")

        

        current_row += 1

    wd.save(READY_DIR / "Списки_поступающих_бакалавриат_баллы.xlsx")

    print("\nЗакончил работать со вступительными списками.")
    print("-----------------------------------------------")

    # проверяем стдуентов, баллы которых не получилось вставить
    print("Список студентов, чьи баллы не удалось вставить. >\n")
    for i in data.keys():
        if i not in inserted_data:
            print(i)






def manager():
    res = None
    # собираем данные из ведомостей
    res = collect_data_statement(log_mode=False)


    if not res:
        with open(CACHE_DIR / "json.json", mode="r", encoding="utf-8") as f:
            res = json.load(f)

    admission_lists_add_points(res, False)



if __name__ == "__main__":
    manager()
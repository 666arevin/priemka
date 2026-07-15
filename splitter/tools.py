import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter

# Настройка стилей
thin_side = openpyxl.styles.Side(border_style="thin", color="000000")
my_border = openpyxl.styles.Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
my_font = openpyxl.styles.Font(name="Times New Roman", size=10)
my_alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="center", horizontal="center")

pattern: str = "Спортивная подготовка по виду спорта. Тренерско-преподавательская деятельность в образовании"

def insert_data_to_row(ws: Worksheet, row_out: list | tuple[Cell, ...], row_id: int,
                        styles: bool = False, merged_cells: list = None, numeric: int = None) -> None:
    """Функция вставляет данные в определенную строку

    Args:
        ws (WorkSheet): рабочий лист WoorkBook
        row_out (list | tuple): список ячеек, которые нужно вставить 
        row_id (int): id строки, куда надо вставить ячейки
        styles (bool): флаг, указывающий, нужно ли применять стили
        merged_cells (list): список координат смерженных ячеек
    """
    # получаем строку, куда будем вставлять данные
    row_in = ws[row_id]

    # проходим циклом по ячейкам в строке и вставляем данные
    for cell_in, cell_out in zip(row_in, row_out):
        cell_in.value = cell_out.value

        # применяем стили, если флаг styles установлен в True
        if styles:
            cell_in.border = my_border
            cell_in.font = my_font
            cell_in.alignment = my_alignment
            
            # мержим некоторые колонки 
            for letters in merged_cells:
                band = f"{letters[0]}{row_id}:{letters[1]}{row_id}"
                ws.merge_cells(band)

            # специальные стили для некоторых ячеек
            if isinstance(cell_out.value, str):
                cell_text = cell_out.value.lower()
            else:
                continue

            if cell_text in ["заочная", "очная"]:
                cell_in.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="center")

            elif cell_text.startswith(pattern):
                cell_in.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="center")
            
            # настраиваем высоту
            ws.row_dimensions[row_id].height = 160
    
    ws[row_id][0].value = numeric


def safe_insert_rows(ws: Worksheet, start_row: int, amount: int):
    """Безопасно вставляем строки сдвигая смерженные ячейки.

    Args:
        ws (Worksheet): рабочий лист
        start_row (int): строка, с которой начинается вставка
        amount (int): количество строк для вставки
    """

    # 1. Сначала сдвигаем границы всех объединенных диапазонов, которые находятся ниже
    for merged_range in list(ws.merged_cells.ranges):

        # Если объединенная ячейка начинается ниже или на уровне вставки
        if merged_range.min_row >= start_row:
            merged_range.shift(row_shift=amount, col_shift=0)

        # Если вставка происходит ВНУТРИ объединенного диапазона (расширяем его)
        elif merged_range.max_row >= start_row:
            merged_range.max_row += amount
        

    # 2. Теперь выполняем стандартную вставку строк
    ws.insert_rows(idx=start_row, amount=amount)
    
    

def formatted_merged_cells(ws):
    for merged_range in ws.merged_cells.ranges:
        columns_count = merged_range.max_col - merged_range.min_col + 1

        if columns_count >= 18:
            row_idx = merged_range.min_row
            ws.row_dimensions[row_idx].height = 13

def height_formatted(ws: Worksheet) -> None:
    pattern_uuid: str = "уникальный код"
    pattern_profile: str = "профиль"

    profile_id = -1

    for row_idx, row in enumerate(ws.iter_rows(), 1):

        if row[0].value == "№":
            # получаем информацию о смерженных ячейках и id столбцов
            uuid_dict, merged_cells = read_headers(row, pattern_uuid, pattern_profile)
            # получаем индексы нужных нам колонок
            profile_id = uuid_dict[pattern_profile]

        # если в стлобце вида спорта, есть паттерн, настриваем высоту и мержим ячейки
        if isinstance(row[profile_id].value, str):
            ws.row_dimensions[row_idx].height = 160
            
            # мержим ячейки
            for letters in merged_cells:
                band = f"{letters[0]}{row_idx}:{letters[1]}{row_idx}"
                ws.merge_cells(band)

        # если пустая строка, выставляем высоту 13
        elif row[profile_id].value is None and row[0].value is None:
            ws.row_dimensions[row_idx].height = 13


def read_headers(row: tuple[Cell, ...], *args) -> tuple[dict, list]:
    """Функция ищет в строке необходимую ячейку и возвращает ее индекс,
    также находит в строке все merge ячейки и возвращает их координаты.

    Args:
        row (tuple[Cell, ...]): кортеж из клеток
        args: можно передать нограниченное количество столбцов, которые
        необходимо найти. Названия столбцов должно быть в нижнем регистре.

    Returns:
        tuple[dict, list]: возврщает id столбцов и координаты смерженных ячеек.
    """
    
    columns_id = dict()
    merged_sells = list()

    # указатели для определения смерженных ячеек
    up_left_corner = None
    down_right_corner = None

    log = list()
    
    # проходим циклом по ячейкам в строчке
    for id, cell in enumerate(row):

        log.append(cell.value)
        # если встретилась первая в строчке смерженная клетка, запоминаем
        if isinstance(cell, MergedCell) and up_left_corner is None:
            column_num = row[id - 1].column
            up_left_corner = get_column_letter(column_num)

        # если встретилась обычная клетка, запоминаем правую границу
        elif isinstance(cell, Cell) and up_left_corner:
            column_num = row[id - 1].column
            down_right_corner =get_column_letter(column_num)
        
        # записываем найденные координаты
        if up_left_corner and down_right_corner:
            merged_sells.append((up_left_corner, down_right_corner))

            # обнуляем указатели
            up_left_corner = None
            down_right_corner = None

        # находим нужные колонки и возвращаем их column_id
        if isinstance(cell.value, str) and cell.value.lower() in args:
            cell_text_lower = cell.value.lower()
            columns_id[cell_text_lower] = id

    # елси после цикла остался up_left_corner, значит был merge
    if up_left_corner:
        column_num = cell.column
        down_right_corner = get_column_letter(column_num)
        merged_sells.append((up_left_corner, down_right_corner))

    # проверка что все столбцы были найдены
    if len(columns_id) < len(args):

        # выводим все колонки, которые не были надены
        ex_list = list()
        for i in args:
            if i not in columns_id.keys():
                ex_list.append(i)
        raise Exception(f"Не удалось найти колонки - > {ex_list}, колонки в строке -> {log}")
    

    return (columns_id, merged_sells)
            
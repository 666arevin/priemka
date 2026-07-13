import openpyxl
import re
from tools import *
from pathlib import Path
import json
import os
from datetime import datetime

BASE_DIR = Path(__file__).parent.resolve()
READY_DIR = BASE_DIR / "ready"
SOURCE_DIR = BASE_DIR / "source"


# ищем по регулярке файл
def find_file_by_regex(folder_path, pattern):
    # Компилируем регулярное выражение для скорости
    regex = re.compile(pattern)
    
    # Перебираем все файлы в указанной папке
    for filename in os.listdir(folder_path):
        # Проверяем, подходит ли имя файла под регулярку
        if regex.match(filename):
            # Возвращаем полный путь к файлу
            return os.path.join(folder_path, filename)
    
    raise FileNotFoundError("Файл не найден")

path_gosuslugi = find_file_by_regex(SOURCE_DIR, r"^Все_заявления.*\.xlsx$")
path_admission = Path(find_file_by_regex(SOURCE_DIR, r"^Списки поступающих бакалавриат.*\.xlsx$"))



# Получаем текущую дату и время
now = datetime.now()
# Выводим в стандартном формате (ГГГГ-ММ-ДД ЧЧ:ММ:СС)
now = now.strftime("%Y-%m-%d %H:%M:%S")

# переменная для записи логов
logs = "\n\n" + "Программа была запущенна в " + now + "\n\n"


def gosuslugi_splitter(no_editing: bool, log_mode: bool):
    """Функция обрабатывает файл с госуслуг, формируя на его основе json файл
    с видами спорта, которые нужно добавить в файл с 1С (списки поступащих).

    Returns:
        dict: Возвращает словарик с данными о видах спорта.
    """
    print("\nРаботаю с файлом с госуслуг.\n")

    # загружаем таблицу excel и берем нудный лист
    wd = openpyxl.load_workbook(path_gosuslugi)
    ws = wd["Sheet1"]

    global logs
    data: dict = {} # словарик для кеша данных из excel gosuslugi
    current_row: int = 1

    if no_editing:
        # получаем словарь уже записанных студентов
        with open(BASE_DIR / "json.json", mode="r", encoding="utf-8") as f:
            data = json.load(f)


    # проходимся по всем строкам таблицы
    while current_row <= ws.max_row:
        sports = None
        row = ws[current_row]
        uuid = str(row[1].value)

        # если включен спец режим без редактирования
        if no_editing:

            # деаем правильную запись новых видов спорта list или str
            if re.search(r"\s*,\s*", str(row[34].value)):
                sports = re.split(pattern=r"\s*,\s*", string=row[34].value)
                sports = list(set(sports))
            else:
                sports = row[34].value

            # если пользователя еще нет в кэше, добавляем его
            if uuid not in data and row[34].value:
                # сохраняем результат в словарик
                data[uuid] = sports

        
        # ищем в какой строке в столбце 34 есть запятая
        elif re.search(r"\s*,\s*", str(row[34].value)):

            # получаем номер строки куда будет происзодить вставка новых строк
            # берем сразу следующую строку
            id = row[34].row + 1

            # получаем список всех видов спорта деля по запятой
            sports = re.split(pattern=r"\s*,\s*", string=row[34].value)
            sports = list(set(sports))
            # получаем количество строк, которое нужно добавить, -1 так как одна уже есть
            amount_split = len(sports) - 1

            # сохраняем виды спорта, которые нужно добавить, а также индентификатор
            data[str(row[1].value)] = sports
            
            # далее начинаем изменять таблицу
            # уже существующую строку изменяем, вписывая один вид спорта
            row[34].value = sports[0]

            if amount_split > 0: # специальная проверка, если в списке больше 1 вида спорта
                # вставляем пустую строку двигая другие строки
                ws.insert_rows(idx=id, amount=amount_split)
                current_row += amount_split

            # копируем строку и вставляем новые данные
            for sport in sports[1:]:
                insert_data_to_row(ws, row, id)

                # вставляем нужный вид спорта и переходим на следующую строку
                ws[id][34].value = sport
                id += 1

        # условие если в табличке 1 вид спорта без запятой
        elif row[34].value and row[1].value not in data:
            sports = row[34].value
            data[str(row[1].value)] = sports
        
        # увеличиваем счетчик строки
        current_row += 1

        # Логирование
        if log_mode and row[34].value:
            print(f"uuid -> {uuid}, спорт -> {sports}")

        # записываем как отработал цикл
        logs += f"Строка -> {current_row}, вид или виды спорта -> {sports}" + "\n"

    try:
        wd.save(BASE_DIR / "temp" / "gosuslugi.xlsx")
        print("\nФайл с госуслуг готов!\n")
    except Exception as e:
        print(f"\nПерехватил ошибку -> {e}\n")
    finally:
        # сохраняем логи
        with open(BASE_DIR / "logs.txt", mode="w", encoding="utf-8") as f:
            f.write(logs)

    
    return data




def admission_lists_splitter_add(data: dict):
    """Функция редактирует и форматирует файл с 1C (вступительные списки)

    Args:
        data (dict): словарь с данными, которые нужно вставить.
    """
    print("\nРаботаю с файлом вступительных списков\n")

    # загружаем файл excel
    wd_admission = openpyxl.load_workbook(path_admission)
    ws_adm = wd_admission.active

    pattern: str = "Спортивная подготовка по виду спорта. Тренерско-преподавательская деятельность в образовании"
    pattern_uuid: str = "уникальный код"
    pattern_profile: str = "профиль"
    current_row: int = 700
    
    # нумерация и id
    profile_id = -1
    numeric = 0

    missed_gosuslugi = list()
    missed_admission = list()

    # проходимся по всем строкам таблицы    
    while current_row <= ws_adm.max_row:
        row = ws_adm[current_row]

        # если мы наткнулись на заголок таблицы, получаем индексы столбцов
        if row[0].value == "№":
            uuid_dict, merged_cells = read_headers(row, pattern_uuid, pattern_profile)

            # получаем индексы нужных нам колонок
            uuid_id = uuid_dict[pattern_uuid]
            profile_id = uuid_dict[pattern_profile]

            # сбрасываем нумерацию
            numeric = 0

        # если нашли нужную строку с видом спорта
        if row[profile_id].value == pattern:

            # проверка на назождение столбца для вставки данных
            if profile_id == -1:
                raise Exception("Не был найден столбец для вставки данных.")

            cell_text = str(row[uuid_id].value)
            sports = data.get(cell_text)
            missed_admission.append(cell_text)

            # есть в вступительный списках нет на госуслугах
            if sports is None:
                missed_gosuslugi.append(cell_text)
            #     print(f"uuid_id -> {str(row[uuid_id].value)}")
            #     input("Продолжить?")

            # производим корректную нумеровку
            row[0].value = numeric
            print(f"row: {current_row}, sports: {sports}")

            # проверяем сколько видов спорта надо вставить, если str то один
            if isinstance(sports, str):

                # добавляем вид спорта и форматруем высоту строки
                row[profile_id].value += "; " + sports
                ws_adm.row_dimensions[current_row].height = 160

            # если список, то нужно вставить несколько строк
            elif isinstance(sports, list):
                # изменяем текущию строку
                row[profile_id].value += "; " + sports[0]

                # получаем количество строк которое нужно вставить
                amount_split = len(sports) - 1

                # важная проверка, если в списке только один обьект дальше не нужно ничего делать
                if amount_split <= 0:
                    continue

                # берем следующую строку, так как текущая уже заполнена и вставляем новые строки
                id = current_row + 1
                safe_insert_rows(ws_adm, id, amount_split)    

                # заполняем новые строки
                for sport in sports[1:]:

                    # вставляем строки применяя к ним стиль
                    insert_data_to_row(
                        ws_adm,
                        row_out=row,
                        row_id=id, 
                        styles=True,
                        merged_cells=merged_cells,
                        numeric=numeric,
                        )
                    
                    # добаялем вид спорта
                    ws_adm[id][profile_id].value = pattern + "; " + sport
                    id += 1
        
                current_row += amount_split
            else:
                print(f"Тип данных -> {type(sports)}, ключ -> {row[1].value}, тип ключа -> {type(row[1].value)}")
                print(f"ДЛина словаря -> {len(data)}")
            

        numeric += 1      
        current_row += 1

    # форматируем все обьединенные ячейки
    print("Форматирую высоту merge клеток")
    formatted_merged_cells(ws_adm)

    # форматируем высоту клеток
    print("Форматирую высоту клеток и делаю merge")
    height_formatted(ws_adm)

    # wd_admission.save(BASE_DIR / "ready" / "excel.xlsx")
    wd_admission.save(BASE_DIR / "ready" / "Списки_поступающих_обработанные.xlsx")

    print("\nЗакончил работу с файлом вступительных списков\n")

    return missed_gosuslugi, missed_admission




def strart_programm(no_editing: bool = False, use_cashe: bool = False,
                    log_mode: bool = False):
    """Функция менеджер, которая запускает программу и управляет ей.

    Args:
        no_editing (bool[True, False]): Программа gosuslugi_splitter не будет изменять файл excel,
                                        лишь только соберет нужные данные и завершит свою работу
        use_cashe (bool[True, False]): Не использется
        use_cashe (bool[True, False]): Активация вывода логов в терминал во время работы программы.
    """
    res = None
    # обрабатываем файл с госуслуг
    res = gosuslugi_splitter(no_editing, log_mode)

    with open(BASE_DIR / "json.json", mode="w", encoding="utf-8") as f:
        json.dump(res, f, ensure_ascii=False, indent=4)
    
    # если не была выполнена функция gosuslugi_splitter, берем данные из кэша
    if not res:
        with open(BASE_DIR / "json.json", mode="r", encoding="utf-8") as f:
            res = json.load(f)

    # missed_gosuslugi id которые не были найдены на госуслугах
    # all_founded все id которые были найдены в вступительных списках
    missed_gosuslugi, all_founded = admission_lists_splitter_add(res)

    missed_admission = list()
    # проверяем каких нет во вступительных списках
    for i in res.keys():
        if i not in all_founded:
            missed_admission.append(i)

    expression1 = "не были найдены на госуслугах, но есть в вступительных списках."
    expression2 = "не был найден в вступительных списках, но есть на госуслугах."

    # сохраняем не найденные id
    with open(BASE_DIR / "errors" / "gosuslugi.txt", mode="w", encoding="utf-8") as f:
        missed_gosuslugi = expression1 + "\n" + str(missed_gosuslugi)
        f.write(str(missed_gosuslugi))

    with open(BASE_DIR / "errors" / "admission.txt", mode="w", encoding="utf-8") as f:
        missed_admission = expression2 + "\n" + str(missed_admission)
        f.write(str(missed_admission))


    print("ID всех не найденны студентов храняться в папке errors.")
    print(f"Файл gosuslugi -> {expression1}")
    print(f"Файл admission -> {expression2}")


if __name__ == "__main__":
    strart_programm(no_editing=False, log_mode=True)



